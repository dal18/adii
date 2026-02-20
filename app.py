import os
import sqlite3
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from dotenv import load_dotenv
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from database import get_db, init_db
from decorators import admin_required
import tempfile

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')

# Admin credentials from environment
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')

# Initialize database on startup (with error handling)
try:
    with app.app_context():
        init_db()
        print("Database initialized successfully.")
except Exception as e:
    print(f"Database initialization failed: {e}")
    # Continue without database for now

# ============================================================================
# PUBLIC ROUTES (Viewer access - no login required)
# ============================================================================

@app.route('/', methods=['GET', 'POST'])
def home():
    """Home page with logbook entry form (public)"""
    if request.method == 'POST':
        try:
            conn = get_db()
            cursor = conn.cursor()
            
            # Handle position: if "None" selected, use custom_position
            position = request.form.get('position', '')
            if position == 'None':
                position = request.form.get('custom_position', '')
            
            # Handle purpose: if "Others" selected, use custom_purpose
            purpose = request.form.get('purpose', '')
            if purpose == 'Others':
                purpose = request.form.get('custom_purpose', '')
            
            # Handle office: if "None" selected, use custom_office
            office = request.form.get('office', '')
            if office == 'None':
                office = request.form.get('custom_office', '')
            
            cursor.execute('''
                INSERT INTO logbook (date, time, name, gender, age, purpose, position, contact_number, address, office, signature)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                request.form['date'],
                request.form['time'],
                request.form['name'],
                request.form['gender'],
                int(request.form['age']),
                purpose,
                position,
                request.form['contact_number'],
                request.form.get('address', ''),
                office,
                request.form['signature']
            ))
            
            conn.commit()
            conn.close()
            
            flash('Thank you! Your entry has been submitted successfully.', 'success')
            return redirect(url_for('home'))
            
        except Exception as e:
            flash(f'Error submitting entry: {str(e)}', 'danger')
            return redirect(url_for('home'))
    
    return render_template('home.html')

# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login page"""
    # If already logged in, redirect to admin dashboard
    if session.get('is_admin'):
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['is_admin'] = True
            session['admin_username'] = username
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'danger')
    
    return render_template('login.html')

@app.route('/logout', methods=['POST'])
def logout():
    """Logout admin and redirect to home"""
    session.clear()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('home'))

# ============================================================================
# ADMIN ROUTES (protected with @admin_required)
# ============================================================================

@app.route('/admin')
@admin_required
def dashboard():
    """Admin dashboard with stats and recent entries"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get today's date
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Total entries
    cursor.execute('SELECT COUNT(*) FROM logbook')
    total_entries = cursor.fetchone()[0]
    
    # Today's entries
    cursor.execute('SELECT COUNT(*) FROM logbook WHERE date = ?', (today,))
    today_entries = cursor.fetchone()[0]
    
    # Gender breakdown
    cursor.execute('SELECT gender, COUNT(*) as count FROM logbook GROUP BY gender')
    gender_stats = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Recent 8 entries
    cursor.execute('''
        SELECT * FROM logbook 
        ORDER BY created_at DESC 
        LIMIT 8
    ''')
    recent_entries = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('admin/dashboard.html',
                         total_entries=total_entries,
                         today_entries=today_entries,
                         gender_stats=gender_stats,
                         recent_entries=recent_entries)

@app.route('/admin/logbook')
@admin_required
def admin_logbook_index():
    """View all logbook entries in DataTable"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Check if filtering by today
    filter_today = request.args.get('filter') == 'today'
    today = datetime.now().strftime('%Y-%m-%d')
    
    if filter_today:
        cursor.execute('SELECT * FROM logbook WHERE date = ? ORDER BY created_at DESC', (today,))
    else:
        cursor.execute('SELECT * FROM logbook ORDER BY created_at DESC')
    
    entries = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('admin/logbook/index.html', entries=entries, filter_today=filter_today)

@app.route('/admin/logbook/<int:id>')
@admin_required
def admin_logbook_show(id):
    """View single logbook entry detail"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM logbook WHERE id = ?', (id,))
    entry = cursor.fetchone()
    
    conn.close()
    
    if not entry:
        flash('Entry not found.', 'danger')
        return redirect(url_for('admin_logbook_index'))
    
    return render_template('admin/logbook/show.html', entry=dict(entry))

@app.route('/admin/logbook/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_logbook_edit(id):
    """Edit logbook entry"""
    conn = get_db()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        try:
            # Handle position: if "None" selected, use custom_position
            position = request.form.get('position', '')
            if position == 'None':
                position = request.form.get('custom_position', '')
            
            # Handle purpose: if "Others" selected, use custom_purpose
            purpose = request.form.get('purpose', '')
            if purpose == 'Others':
                purpose = request.form.get('custom_purpose', '')
            
            # Handle office: if "None" selected, use custom_office
            office = request.form.get('office', '')
            if office == 'None':
                office = request.form.get('custom_office', '')
            
            cursor.execute('''
                UPDATE logbook 
                SET date = ?, time = ?, name = ?, gender = ?, age = ?, 
                    purpose = ?, position = ?, contact_number = ?, address = ?, office = ?, signature = ?
                WHERE id = ?
            ''', (
                request.form['date'],
                request.form['time'],
                request.form['name'],
                request.form['gender'],
                int(request.form['age']),
                purpose,
                position,
                request.form['contact_number'],
                request.form.get('address', ''),
                office,
                request.form['signature'],
                id
            ))
            
            conn.commit()
            conn.close()
            
            flash('Entry updated successfully!', 'success')
            return redirect(url_for('admin_logbook_show', id=id))
            
        except Exception as e:
            conn.close()
            flash(f'Error updating entry: {str(e)}', 'danger')
            return redirect(url_for('admin_logbook_edit', id=id))
    
    # GET request - show edit form
    cursor.execute('SELECT * FROM logbook WHERE id = ?', (id,))
    entry = cursor.fetchone()
    conn.close()
    
    if not entry:
        flash('Entry not found.', 'danger')
        return redirect(url_for('admin_logbook_index'))
    
    return render_template('admin/logbook/edit.html', entry=dict(entry))

def export_to_excel():
    """Export logbook entries to Excel with signatures"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Check if filtering by today
    filter_today = request.args.get('filter') == 'today'
    today = datetime.now().strftime('%Y-%m-%d')
    
    if filter_today:
        cursor.execute('SELECT * FROM logbook WHERE date = ? ORDER BY created_at DESC', (today,))
        filename = f'logbook_today_{today}.xlsx'
    else:
        cursor.execute('SELECT * FROM logbook ORDER BY created_at DESC')
        filename = f'logbook_all_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    
    entries = cursor.fetchall()
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Logbook Entries"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    column_widths = {
        'A': 8,   # ID
        'B': 12,  # Date
        'C': 10,  # Time
        'D': 25,  # Name
        'E': 10,  # Gender
        'F': 8,   # Age
        'G': 20,  # Position
        'H': 15,  # Purpose
        'I': 18,  # Contact
        'J': 25,  # Address
        'K': 15,  # Office
        'L': 20,  # Signature
        'M': 20   # Created At
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Create headers
    headers = ['ID', 'Date', 'Time', 'Name', 'Gender', 'Age', 'Position', 'Purpose', 'Contact Number', 'Address', 'Office', 'Signature', 'Created At']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    temp_files = []  # Keep track of temp files to clean up later
    
    for row_num, entry in enumerate(entries, 2):
        # Basic data
        ws.cell(row=row_num, column=1, value=entry['id'])
        ws.cell(row=row_num, column=2, value=entry['date'])
        ws.cell(row=row_num, column=3, value=entry['time'])
        ws.cell(row=row_num, column=4, value=entry['name'])
        ws.cell(row=row_num, column=5, value=entry['gender'])
        ws.cell(row=row_num, column=6, value=entry['age'])
        ws.cell(row=row_num, column=7, value=entry['position'])
        ws.cell(row=row_num, column=8, value=entry['purpose'])
        ws.cell(row=row_num, column=9, value=entry['contact_number'])
        ws.cell(row=row_num, column=10, value=entry['address'] or '')
        ws.cell(row=row_num, column=11, value=entry['office'] or '')
        ws.cell(row=row_num, column=13, value=entry['created_at'])
        
        # Handle signature - add actual signature image to Excel
        if entry['signature']:
            try:
                # Decode base64 signature
                signature_data = entry['signature'].split(',')[1] if ',' in entry['signature'] else entry['signature']
                signature_bytes = base64.b64decode(signature_data)
                
                # Create a temporary file for signature image
                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_sig:
                    temp_sig.write(signature_bytes)
                    temp_sig.flush()  # Ensure data is written
                    temp_sig_path = temp_sig.name
                    temp_files.append(temp_sig_path)  # Add to cleanup list
                
                # Add signature image to Excel with proper cell positioning
                img = Image(temp_sig_path)
                img.width = 80
                img.height = 40
                
                # Position image within signature cell (column L)
                # Simple string anchor for positioning
                img.anchor = f'L{row_num}'
                
                ws.add_image(img)
                
                # Adjust row height for signature
                ws.row_dimensions[row_num].height = 60
                
            except Exception as e:
                print(f"Error processing signature for entry {entry['id']}: {e}")
                ws.cell(row=row_num, column=12, value="[Signature Error]")
        else:
            ws.cell(row=row_num, column=12, value="[No Signature]")
        
        # Apply borders to data cells
        for col_num in range(1, 14):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Add summary section at the top
    ws.insert_rows(1)
    ws.insert_rows(1)
    
    # Summary title
    title_cell = ws.cell(row=1, column=1, value=f"LOGBOOK EXPORT - {'TODAY' if filter_today else 'ALL ENTRIES'}")
    title_cell.font = Font(name='Arial', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='left')
    ws.merge_cells(f'A1:K1')
    
    # Summary info
    export_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    ws.cell(row=2, column=1, value=f"Exported on: {export_date}")
    ws.cell(row=2, column=1).font = Font(name='Arial', size=11, italic=True)
    ws.merge_cells(f'A2:K2')
    
    # Statistics
    total_entries = len(entries)
    male_count = sum(1 for entry in entries if entry['gender'] == 'Male')
    female_count = sum(1 for entry in entries if entry['gender'] == 'Female')
    
    ws.cell(row=3, column=1, value=f"Total Entries: {total_entries}")
    ws.cell(row=4, column=1, value=f"Male: {male_count} | Female: {female_count}")
    ws.cell(row=3, column=1).font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=4, column=1).font = Font(name='Arial', size=11)
    
    # Add note about signatures
    ws.cell(row=5, column=1, value="Note: Signatures are stored in the database and can be viewed in the admin interface.")
    ws.cell(row=5, column=1).font = Font(name='Arial', size=10, italic=True)
    ws.merge_cells(f'A5:K5')
    
    # Save to memory
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Clean up temporary files
    for temp_file in temp_files:
        try:
            os.unlink(temp_file)
        except:
            pass  # Ignore cleanup errors
    
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/export/excel')
@admin_required
def admin_export_excel():
    """Admin route to export logbook to Excel"""
    return export_to_excel()

@app.route('/admin/logbook/<int:id>/delete', methods=['POST'])
@admin_required
def admin_logbook_delete(id):
    """Delete logbook entry"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('DELETE FROM logbook WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        
        flash('Entry deleted successfully!', 'success')
    except Exception as e:
        conn.close()
        flash(f'Error deleting entry: {str(e)}', 'danger')
    
    return redirect(url_for('admin_logbook_index'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

# For Railway/Production deployment
# Railway will use the Procfile to start the app
# The app is already properly configured for WSGI
